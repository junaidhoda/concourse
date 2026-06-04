#!/usr/bin/env python3
"""
Airport restaurant scraper — Playwright + openpyxl.

Dependencies (pip):
  playwright beautifulsoup4 lxml openpyxl tenacity

After pip install playwright, install browsers once:
  playwright install chromium

Set HEADLESS = False below to watch the browser while debugging.
Set USE_MOCK_DATA = True to exercise Excel export only (no network).
"""

from __future__ import annotations

import logging
import re
import sys
import time
from datetime import datetime, timezone
from typing import Any
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from playwright.sync_api import Page, sync_playwright, TimeoutError as PlaywrightTimeout
from tenacity import retry, stop_after_attempt, wait_exponential

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

HEADLESS = True
USE_MOCK_DATA = False

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

PAGE_TIMEOUT_MS = 30_000
INTER_REQUEST_DELAY_S = 2.0

# Official display names for summary / rows
AIRPORT_NAMES: dict[str, str] = {
    "BHX": "Birmingham Airport",
    "BKK": "Suvarnabhumi Airport (Bangkok)",
    "CDG": "Paris Charles de Gaulle",
    "DXB": "Dubai International",
    "FRA": "Frankfurt Airport",
    "IST": "Istanbul Airport",
    "JFK": "John F. Kennedy International",
    "LAX": "Los Angeles International",
    "LGW": "London Gatwick",
    "LHR": "London Heathrow",
    "MAN": "Manchester Airport",
    "SIN": "Singapore Changi",
}

# Seed URL per airport (entry points)
AIRPORT_URLS: dict[str, str] = {
    "BHX": "https://www.birminghamairport.co.uk/at-the-airport/shopping-and-dining/",
    "BKK": "https://suvarnabhumi.airportthai.co.th/en/category/concession/restaurant/",
    "CDG": "https://www.parisaeroport.fr/en/passengers/services/shops-and-restaurants",
    "DXB": "https://dubaiairports.ae/experiences/restaurants",
    "FRA": "https://www.frankfurt-airport.com/en/shopping---dining.html",
    "IST": "https://www.igairport.com/en/shopping-and-food/food-and-beverages",
    "JFK": "https://www.jfkairport.com/dine",
    "LAX": "https://www.laxshopdine.com/directory",
    "LGW": "https://www.gatwickairport.com/at-the-airport/eat-drink-shop/",
    "LHR": "https://www.heathrow.com/at-the-airport/restaurants-a-z",
    "MAN": "https://www.manchesterairport.co.uk/eat-shop-relax/eating-and-drinking/",
    "SIN": "https://www.changiairport.com/en/dine.html",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(levelname)s %(message)s",
    stream=sys.stdout,
)


# ---------------------------------------------------------------------------
# Schema helpers
# ---------------------------------------------------------------------------


def scraped_at_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def row_dict(
    airport_code: str,
    airport_name: str,
    restaurant_name: str,
    terminal: str | None,
    floor: str | None,
    cuisine: str | Nonea,
    location_detail: str | None,
    source_url: str,
) -> dict[str, Any]:
    return {
        "airport_code": airport_code,
        "airport_name": airport_name,
        "restaurant_name": restaurant_name,
        "terminal": terminal,
        "floor": floor,
        "cuisine": cuisine,
        "location_detail": location_detail,
        "source_url": source_url,
        "scraped_at": scraped_at_iso(),
    }


COLUMN_ORDER = [
    "airport_code",
    "airport_name",
    "restaurant_name",
    "terminal",
    "floor",
    "cuisine",
    "location_detail",
    "source_url",
    "scraped_at",
]


# ---------------------------------------------------------------------------
# Navigation throttling & retries (per domain 2s minimum between loads)
# ---------------------------------------------------------------------------

_last_nav_time: dict[str, float] = {}


def _domain(url: str) -> str:
    return urlparse(url).netloc or ""


def throttle_same_site(url: str) -> None:
    dom = _domain(url)
    if not dom:
        return
    now = time.monotonic()
    last = _last_nav_time.get(dom)
    if last is not None:
        wait = INTER_REQUEST_DELAY_S - (now - last)
        if wait > 0:
            time.sleep(wait)
    _last_nav_time[dom] = time.monotonic()


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=20),
    reraise=True,
)
def goto_with_retry(page: Page, url: str) -> None:
    """Navigate with 3 attempts and exponential backoff (tenacity)."""
    throttle_same_site(url)
    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
    try:
        page.wait_for_load_state("networkidle", timeout=15_000)
    except PlaywrightTimeout:
        pass


# ---------------------------------------------------------------------------
# Shared UI helpers
# ---------------------------------------------------------------------------


def dismiss_cookie_banners(page: Page) -> None:
    selectors = [
        "#onetrust-accept-btn-handler",
        'button:has-text("Accept")',
        'button:has-text("Accept all")',
        'button:has-text("I agree")',
        'button:has-text("Allow all")',
        'button[aria-label="Accept"]',
        'a.cc-btn.cc-dismiss',  # common cookie consent
    ]
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            if loc.count() > 0 and loc.is_visible(timeout=2000):
                loc.click(timeout=3000)
                time.sleep(0.5)
        except Exception:
            continue


def scroll_to_bottom_until_stable(page: Page, rounds: int = 25, pause_ms: int = 800) -> None:
    last_h = -1
    stable = 0
    for _ in range(rounds):
        h = page.evaluate("() => document.body ? document.body.scrollHeight : 0")
        if h == last_h:
            stable += 1
            if stable >= 2:
                break
        else:
            stable = 0
        last_h = h
        page.evaluate("() => window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(pause_ms)


def click_load_more_until_gone(page: Page, max_clicks: int = 80) -> None:
    texts = ["Load more", "Show more", "View more", "See more", "Load More"]
    for _ in range(max_clicks):
        clicked = False
        for t in texts:
            try:
                btn = page.get_by_role("button", name=re.compile(re.escape(t), re.I)).first
                if btn.count() > 0 and btn.is_visible(timeout=1000):
                    btn.click(timeout=5000)
                    clicked = True
                    page.wait_for_timeout(1500)
                    break
            except Exception:
                continue
        if not clicked:
            try:
                link = page.locator(f'a:has-text("{texts[0]}")').first
                if link.count() > 0 and link.is_visible(timeout=500):
                    link.click(timeout=5000)
                    page.wait_for_timeout(1500)
                    clicked = True
            except Exception:
                pass
        if not clicked:
            break


def soup_from_page(page: Page) -> BeautifulSoup:
    return BeautifulSoup(page.content(), "lxml")


# ---------------------------------------------------------------------------
# Mock data (Excel pipeline test)
# ---------------------------------------------------------------------------


def mock_rows_for_airport(code: str) -> list[dict[str, Any]]:
    name = AIRPORT_NAMES[code]
    url = AIRPORT_URLS[code]
    ts = scraped_at_iso()
    return [
        {
            "airport_code": code,
            "airport_name": name,
            "restaurant_name": f"[MOCK] Sample A at {code}",
            "terminal": None,
            "floor": None,
            "cuisine": None,
            "location_detail": None,
            "source_url": url,
            "scraped_at": ts,
        },
        {
            "airport_code": code,
            "airport_name": name,
            "restaurant_name": f"[MOCK] Sample B at {code}",
            "terminal": "T1",
            "floor": None,
            "cuisine": "International",
            "location_detail": "Airside",
            "source_url": url,
            "scraped_at": ts,
        },
    ]


# ==============================================================================
# Airport scrapers (one per airport)
# ==============================================================================


def scrape_sin(page: Page) -> list[dict[str, Any]]:
    """Changi — terminal tabs + scroll / load more. # TODO: verify selectors after site updates."""

    code, aname = "SIN", AIRPORT_NAMES["SIN"]

    def parse_listing_cards(soup: BeautifulSoup, term_code: str | None) -> list[dict[str, Any]]:
        found: list[dict[str, Any]] = []
        for art in soup.select("article, li[class*='list'], div[class*='card'], div.tile, [class*='shop']"):
            h = art.find(["h2", "h3", "h4"])
            if not h:
                al = art.find("a", href=True)
                if al and len(al.get_text(strip=True)) > 2:
                    h = al
            if not h:
                continue
            name = h.get_text(" ", strip=True)
            if not name or len(name) < 2:
                continue
            nl = name.lower()
            if nl in {"dine", "restaurant", "search", "filter", "learn more", "read more"}:
                continue
            if any(x in nl for x in ["cookie", "privacy", "sign up", "newsletter"]):
                continue
            cuisine = None
            c_el = art.find(class_=re.compile(r"cuisine|category|tag", re.I))
            if c_el:
                cuisine = c_el.get_text(" ", strip=True) or None
            loc = None
            for sel in ["[class*='location']", "[class*='gate']", "[class*='unit']"]:
                le = art.select_one(sel)
                if le:
                    loc = le.get_text(" ", strip=True) or None
                    break
            found.append(
                row_dict(code, aname, name, term_code, None, cuisine, loc, page.url)
            )
        return found

    base = AIRPORT_URLS["SIN"]
    rows: list[dict[str, Any]] = []

    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    time.sleep(1)

    # Try Changi terminal filters (tabs / buttons). Labels vary by site revision.
    terminal_aliases: list[tuple[str, list[str]]] = [
        ("T1", ["Terminal 1", "T1", "Term 1"]),
        ("T2", ["Terminal 2", "T2", "Term 2"]),
        ("T3", ["Terminal 3", "T3", "Term 3"]),
        ("T4", ["Terminal 4", "T4", "Term 4"]),
        ("T5", ["Terminal 5", "T5", "Term 5"]),
    ]

    for term_code, labels in terminal_aliases:
        clicked = False
        for lab in labels:
            for role_name in ("tab", "button", "link"):
                try:
                    loc = page.get_by_role(role_name, name=re.compile(re.escape(lab), re.I)).first
                    if loc.count() > 0 and loc.is_visible(timeout=1200):
                        loc.click(timeout=8000)
                        clicked = True
                        page.wait_for_timeout(2500)
                        break
                except Exception:
                    continue
            if clicked:
                break
        if not clicked:
            logging.info("Changi: no control found for %s — skipping terminal pass", term_code)
            continue

        scroll_to_bottom_until_stable(page)
        click_load_more_until_gone(page)
        soup = soup_from_page(page)
        rows.extend(parse_listing_cards(soup, term_code))
        time.sleep(INTER_REQUEST_DELAY_S)

    # Re-load base once if terminal passes produced nothing (user may land on combined view only)
    if not rows:
        goto_with_retry(page, base)
        dismiss_cookie_banners(page)
        scroll_to_bottom_until_stable(page)
        click_load_more_until_gone(page)
        soup = soup_from_page(page)
        rows.extend(parse_listing_cards(soup, None))

    seen: set[tuple[str, str | None, str]] = set()
    out: list[dict[str, Any]] = []
    for r in rows:
        key = (r["restaurant_name"], r.get("terminal"), r["source_url"])
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def scrape_lhr(page: Page) -> list[dict[str, Any]]:
    """Heathrow — iterate terminal-specific listing pages. # TODO: verify URLs/selectors."""

    code, aname = "LHR", AIRPORT_NAMES["LHR"]
    rows: list[dict[str, Any]] = []

    for term in ["T2", "T3", "T4", "T5"]:
        n = term[1]
        candidates = [
            f"https://www.heathrow.com/at-the-airport/eat-drink-shop/terminal-{n}",
            f"https://www.heathrow.com/at-the-airport/eat-drink-shop/terminal-{n}.html",
            f"https://www.heathrow.com/at-the-airport/eat-drink-shop/t{n}",
            f"https://www.heathrow.com/at-the-airport/eat-drink-shop/t{n}.html",
        ]
        loaded = False
        for turl in candidates:
            try:
                goto_with_retry(page, turl)
                loaded = True
                break
            except Exception:
                continue
        if not loaded:
            logging.warning("LHR: skipped %s — no URL variant loaded", term)
            continue

        dismiss_cookie_banners(page)
        scroll_to_bottom_until_stable(page)
        click_load_more_until_gone(page)
        soup = soup_from_page(page)

        for block in soup.select("article, div[class*='card'], li"):
            title_el = block.find(["h2", "h3", "h4", "strong", "a"])
            if not title_el:
                continue
            name = title_el.get_text(" ", strip=True)
            if not name or len(name) < 2:
                continue
            skip_kw = {"eat", "drink", "shop", "terminal", "heathrow", "filter"}
            if name.lower() in skip_kw:
                continue
            cuisine = None
            p = block.find("p")
            if p:
                tx = p.get_text(" ", strip=True)
                if tx and len(tx) < 200:
                    cuisine = tx
            rows.append(row_dict(code, aname, name, term, None, cuisine, None, page.url))

        time.sleep(INTER_REQUEST_DELAY_S)

    seen: set[tuple[str, str | None]] = set()
    out: list[dict[str, Any]] = []
    for r in rows:
        k = (r["restaurant_name"], r.get("terminal"))
        if k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def scrape_lgw(page: Page) -> list[dict[str, Any]]:
    """Gatwick — North / South filters when present. # TODO: verify selectors."""

    code, aname = "LGW", AIRPORT_NAMES["LGW"]
    url = AIRPORT_URLS["LGW"]
    rows: list[dict[str, Any]] = []

    for term in ("North", "South"):
        goto_with_retry(page, url)
        dismiss_cookie_banners(page)
        try:
            link = page.locator("a, button").filter(has_text=re.compile(rf"\b{term}\b", re.I)).first
            if link.count() > 0:
                link.click(timeout=8000)
                page.wait_for_timeout(2500)
        except Exception:
            pass
        scroll_to_bottom_until_stable(page)
        click_load_more_until_gone(page)
        soup = soup_from_page(page)
        for el in soup.select("article h2, article h3, .card-title, a[href*='eat'], [class*='venue'] h3"):
            name = el.get_text(" ", strip=True)
            if name and 2 < len(name) < 150:
                rows.append(row_dict(code, aname, name, term, None, None, None, page.url))
        time.sleep(INTER_REQUEST_DELAY_S)
    return rows


def scrape_jfk(page: Page) -> list[dict[str, Any]]:
    """JFK dine — terminal/filter iteration. # TODO: verify selectors."""
    code, aname = "JFK", AIRPORT_NAMES["JFK"]
    base = AIRPORT_URLS["JFK"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    rows: list[dict[str, Any]] = []

    for term in ["Terminal 1", "Terminal 4", "Terminal 5", "Terminal 7", "Terminal 8"]:
        try:
            tab = page.get_by_role("tab", name=re.compile(re.escape(term.split()[1]), re.I)).first
            if tab.count() > 0:
                tab.click(timeout=5000)
                page.wait_for_timeout(2000)
        except Exception:
            pass
        scroll_to_bottom_until_stable(page)
        soup = soup_from_page(page)
        for h in soup.select("h2, h3, .title, a.restaurant-name"):
            name = h.get_text(" ", strip=True)
            if name and len(name) > 2:
                rows.append(row_dict(code, aname, name, term.replace("Terminal ", "T"), None, None, None, page.url))
        time.sleep(INTER_REQUEST_DELAY_S)

    if not rows:
        scroll_to_bottom_until_stable(page)
        soup = soup_from_page(page)
        for h in soup.select("h2, h3, li a"):
            name = h.get_text(" ", strip=True)
            if name and 2 < len(name) < 120:
                rows.append(row_dict(code, aname, name, None, None, None, None, base))
    return rows


def scrape_lax(page: Page) -> list[dict[str, Any]]:
    """LAX shop-dine — heavy JS. # TODO: verify selectors."""
    code, aname = "LAX", AIRPORT_NAMES["LAX"]
    base = AIRPORT_URLS["LAX"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    rows: list[dict[str, Any]] = []

    for terminal in ["T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "TBIT"]:
        try:
            btn = page.locator(f'button:has-text("{terminal}"), a:has-text("{terminal}")').first
            if btn.count() > 0 and btn.is_visible(timeout=1500):
                btn.click(timeout=5000)
                page.wait_for_timeout(2000)
        except Exception:
            pass
        scroll_to_bottom_until_stable(page)
        click_load_more_until_gone(page)
        soup = soup_from_page(page)
        for block in soup.select("[class*='venue'], [class*='listing'], article, li"):
            t_el = block.find(["h2", "h3", "h4", "a"])
            if not t_el:
                continue
            name = t_el.get_text(" ", strip=True)
            if name and len(name) > 2:
                rows.append(row_dict(code, aname, name, terminal, None, None, None, page.url))
        time.sleep(INTER_REQUEST_DELAY_S)

    return rows


def scrape_dxb(page: Page) -> list[dict[str, Any]]:
    """DXB dining — load more / scroll. # TODO: verify selectors."""
    code, aname = "DXB", AIRPORT_NAMES["DXB"]
    base = AIRPORT_URLS["DXB"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    click_load_more_until_gone(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, [class*='card'], li"):
        h = el.find(["h2", "h3", "h4"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 2:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_cdg(page: Page) -> list[dict[str, Any]]:
    """CDG shops and restaurants — cookie banner. # TODO: verify selectors."""
    code, aname = "CDG", AIRPORT_NAMES["CDG"]
    base = AIRPORT_URLS["CDG"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for a in soup.select('a[href*="restaurant"], a[href*="food"], .card-title, h2, h3'):
        name = a.get_text(" ", strip=True)
        if name and 3 < len(name) < 200:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_fra(page: Page) -> list[dict[str, Any]]:
    """Frankfurt shopping & dining. # TODO: verify selectors."""
    code, aname = "FRA", AIRPORT_NAMES["FRA"]
    base = AIRPORT_URLS["FRA"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, .teaser, li, [class*='dining']"):
        h = el.find(["h2", "h3", "h4", "a"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 3:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_ist(page: Page) -> list[dict[str, Any]]:
    """Istanbul Airport F&B. # TODO: verify selectors."""
    code, aname = "IST", AIRPORT_NAMES["IST"]
    base = AIRPORT_URLS["IST"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, [class*='card'], [class*='store']"):
        h = el.find(["h2", "h3", "h4", "a"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 2:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_bkk(page: Page) -> list[dict[str, Any]]:
    """Suvarnabhumi concessions (English). # TODO: verify selectors."""
    code, aname = "BKK", AIRPORT_NAMES["BKK"]
    base = AIRPORT_URLS["BKK"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, .card, li, [class*='shop']"):
        h = el.find(["h2", "h3", "h4", "a"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 2:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_bhx(page: Page) -> list[dict[str, Any]]:
    """Birmingham shopping & dining. # TODO: verify selectors."""
    code, aname = "BHX", AIRPORT_NAMES["BHX"]
    base = AIRPORT_URLS["BHX"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, .card, li[class*='listing']"):
        h = el.find(["h2", "h3", "h4"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 2:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


def scrape_man(page: Page) -> list[dict[str, Any]]:
    """Manchester eating & drinking. # TODO: verify selectors."""
    code, aname = "MAN", AIRPORT_NAMES["MAN"]
    base = AIRPORT_URLS["MAN"]
    goto_with_retry(page, base)
    dismiss_cookie_banners(page)
    scroll_to_bottom_until_stable(page)
    rows: list[dict[str, Any]] = []
    soup = soup_from_page(page)
    for el in soup.select("article, .card, [class*='venue']"):
        h = el.find(["h2", "h3", "a"])
        if not h:
            continue
        name = h.get_text(" ", strip=True)
        if name and len(name) > 2:
            rows.append(row_dict(code, aname, name, None, None, None, None, page.url))
    return rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def autofit_columns(ws) -> None:
    for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def write_excel(path: str, rows: list[dict[str, Any]], summary: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws_r = wb.active
    ws_r.title = "Restaurants"
    bold = Font(bold=True)

    for col, key in enumerate(COLUMN_ORDER, start=1):
        c = ws_r.cell(row=1, column=col, value=key)
        c.font = bold
    for i, rec in enumerate(rows, start=2):
        for col, key in enumerate(COLUMN_ORDER, start=1):
            ws_r.cell(row=i, column=col, value=rec.get(key))
    ws_r.freeze_panes = "A2"
    autofit_columns(ws_r)

    ws_s = wb.create_sheet("Summary")
    summary_headers = ["airport_code", "airport_name", "restaurants_found", "scrape_status", "error_message"]
    for col, h in enumerate(summary_headers, start=1):
        c = ws_s.cell(row=1, column=col, value=h)
        c.font = bold
    for i, s in enumerate(summary, start=2):
        ws_s.cell(row=i, column=1, value=s["airport_code"])
        ws_s.cell(row=i, column=2, value=s["airport_name"])
        ws_s.cell(row=i, column=3, value=s["restaurants_found"])
        ws_s.cell(row=i, column=4, value=s["scrape_status"])
        ws_s.cell(row=i, column=5, value=s.get("error_message"))
    ws_s.freeze_panes = "A2"
    autofit_columns(ws_s)

    wb.save(path)


def print_summary_table(summary: list[dict[str, Any]]) -> None:
    lines = [
        "",
        "=== Scrape summary ===",
        f"{'Code':<5} {'Found':<7} {'Status':<9} {'Airport':<26} Error",
        "-" * 110,
    ]
    for s in summary:
        err = (s.get("error_message") or "")[:72]
        nm = (s.get("airport_name") or "")[:24]
        lines.append(
            f"{s['airport_code']:<5} {s['restaurants_found']!s:<7} {s['scrape_status']:<9} {nm:<26} {err}"
        )
    lines.append("=== End summary ===")
    msg = "\n".join(lines)
    print(msg)
    logging.info("%s", msg)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

SCRAPER_FUNCS = [
    ("BHX", scrape_bhx),
    ("BKK", scrape_bkk),
    ("CDG", scrape_cdg),
    ("DXB", scrape_dxb),
    ("FRA", scrape_fra),
    ("IST", scrape_ist),
    ("JFK", scrape_jfk),
    ("LAX", scrape_lax),
    ("LGW", scrape_lgw),
    ("LHR", scrape_lhr),
    ("MAN", scrape_man),
    ("SIN", scrape_sin),
]


def main() -> None:
    out_path = "airport_restaurants.xlsx"
    all_rows: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []

    if USE_MOCK_DATA:
        for code, _fn in SCRAPER_FUNCS:
            mr = mock_rows_for_airport(code)
            all_rows.extend(mr)
            summary.append(
                {
                    "airport_code": code,
                    "airport_name": AIRPORT_NAMES[code],
                    "restaurants_found": len(mr),
                    "scrape_status": "success",
                    "error_message": None,
                }
            )
        write_excel(out_path, all_rows, summary)
        print_summary_table(summary)
        logging.info("Wrote %s (mock data).", out_path)
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(user_agent=USER_AGENT)
        page = context.new_page()
        page.set_default_timeout(PAGE_TIMEOUT_MS)

        for code, fn in SCRAPER_FUNCS:
            aname = AIRPORT_NAMES[code]
            logging.info("Scraping %s...", code)
            try:
                rows = fn(page)
                n = len(rows)
                all_rows.extend(rows)
                logging.info("Scraping %s... found %s restaurants", code, n)
                summary.append(
                    {
                        "airport_code": code,
                        "airport_name": aname,
                        "restaurants_found": n,
                        "scrape_status": "success",
                        "error_message": None,
                    }
                )
            except Exception as e:
                logging.exception("Scraping %s failed: %s", code, e)
                summary.append(
                    {
                        "airport_code": code,
                        "airport_name": aname,
                        "restaurants_found": 0,
                        "scrape_status": "failed",
                        "error_message": str(e),
                    }
                )

        browser.close()

    write_excel(out_path, all_rows, summary)
    print_summary_table(summary)
    logging.info("Wrote %s (%d total rows).", out_path, len(all_rows))


if __name__ == "__main__":
    main()
